#!/usr/bin/env python3

"""
convert_files_to_txt.py

A standalone script that converts various file formats
(.docx, .pdf, .xlsx, .odt, .ipynb, etc.) to plain-text .txt files,
optionally zipping them into one archive at the end.

Usage:
    python convert_files_to_txt.py <input_dir> <output_dir> [--zip <archive_name>]

Example:
    python convert_files_to_txt.py docs/ txt_output/ --zip my_converted_files.zip
"""

import os
import sys
import logging
import json
import unicodedata
import argparse
import zipfile

# Optional parsing libraries
try:
    import docx  # For .docx
except ImportError:
    docx = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import openpyxl  # For .xlsx
except ImportError:
    openpyxl = None

try:
    import odf.text, odf.teletype, odf.opendocument  # For .odt
except ImportError:
    odf = None

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def sanitize_text(text):
    """
    Remove or replace odd/control characters and ensure UTF-8 friendly content.
    - Normalize unicode to NFC form.
    - Remove non-printable control characters except for newlines/tabs.
    """
    text = unicodedata.normalize("NFC", text)
    cleaned_chars = []
    for ch in text:
        if ch.isprintable() or ch in ['\n', '\r', '\t']:
            cleaned_chars.append(ch)
        else:
            cleaned_chars.append(" ")
    sanitized = "".join(cleaned_chars).strip()
    return sanitized


def extract_text_from_file(filepath):
    """
    Attempt to extract text from the given filepath, based on extension.
    Return an empty string if no parser is available or if an error occurs.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (
        ".txt", ".md", ".py", ".json", ".csv", ".tsv", ".log", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart",
        ".php", ".rb", ".sql"
    ):
        return _read_plaintext(filepath)

    if ext == ".docx":
        return _read_docx(filepath) if docx else ""

    if ext in (".doc", ".rtf"):
        logger.warning("Native reading not implemented for %s; consider external tools.", ext)
        return ""

    if ext == ".pdf":
        if fitz:
            return _read_pdf_pymupdf(filepath)
        elif PyPDF2:
            return _read_pdf_pypdf2(filepath)
        else:
            logger.warning("No PDF library installed; cannot parse PDFs.")
            return ""

    if ext == ".odt":
        return _read_odt(filepath) if odf else ""

    if ext in (".xlsx", ".xls", ".xlsm", ".ods"):
        return _read_excel(filepath) if openpyxl else ""

    if ext in (".ppt", ".pptx", ".odp"):
        logger.warning("Parsing for PPT/ODP not implemented.")
        return ""

    if ext == ".ipynb":
        return _read_ipynb(filepath)

    return ""


def _read_plaintext(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        logger.error("Error reading plaintext file %s: %s", filepath, e)
        return ""


def _read_docx(filepath):
    try:
        doc = docx.Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading DOCX %s: %s", filepath, e)
        return ""


def _read_pdf_pymupdf(filepath):
    text_pages = []
    try:
        with fitz.open(filepath) as pdf_doc:
            for page in pdf_doc:
                text_pages.append(page.get_text())
    except Exception as e:
        logger.error("Error reading PDF (PyMuPDF) %s: %s", filepath, e)
    return "\n".join(text_pages)


def _read_pdf_pypdf2(filepath):
    text_pages = []
    try:
        with open(filepath, "rb") as f:
            pdf = PyPDF2.PdfReader(f)
            for page_num in range(len(pdf.pages)):
                text_pages.append(pdf.pages[page_num].extract_text() or "")
    except Exception as e:
        logger.error("Error reading PDF (PyPDF2) %s: %s", filepath, e)
    return "\n".join(text_pages)


def _read_odt(filepath):
    try:
        doc = odf.opendocument.load(filepath)
        text_elements = doc.getElementsByType(odf.text.P)
        paragraphs = [odf.teletype.extractText(elem) for elem in text_elements]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading ODT %s: %s", filepath, e)
        return ""


def _read_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text_chunks = []
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                row_str = "\t".join(str(x) if x is not None else "" for x in row)
                text_chunks.append(row_str)
        return "\n".join(text_chunks)
    except Exception as e:
        logger.error("Error reading Excel file %s: %s", filepath, e)
        return ""


def _read_ipynb(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        text_cells = []
        for cell in data.get("cells", []):
            src = cell.get("source", [])
            text_cells.append("".join(src))
        return "\n".join(text_cells)
    except Exception as e:
        logger.error("Error reading IPYNB %s: %s", filepath, e)
        return ""


def convert_files_to_txt(
    input_dir,
    output_dir,
    max_file_size=5 * 1024 * 1024,
    skip_hidden=True
):
    """
    Recursively scan input_dir for files, extract & sanitize text,
    then write them as .txt in output_dir.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    logger.info("Starting conversion from %s to %s", input_dir, output_dir)

    allowed_extensions = (
        ".txt", ".md", ".py", ".json", ".csv", ".tsv", ".log", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart",
        ".php", ".rb", ".sql", ".doc", ".docx", ".rtf", ".pdf",
        ".odt", ".xls", ".xlsx", ".xlsm", ".ods", ".ppt", ".pptx",
        ".odp", ".ipynb"
    )

    for root, dirs, files in os.walk(input_dir):
        if skip_hidden:
            dirs[:] = [d for d in dirs if not d.startswith(".")]

        for file_name in files:
            if skip_hidden and file_name.startswith("."):
                continue

            ext = os.path.splitext(file_name)[1].lower()
            if ext not in allowed_extensions:
                continue

            full_path = os.path.join(root, file_name)
            # Check file size
            try:
                size = os.path.getsize(full_path)
                if size > max_file_size:
                    logger.debug("Skipping large file %s (size %d bytes)", full_path, size)
                    continue
            except Exception as e:
                logger.error("Could not check file size of %s: %s", full_path, e)
                continue

            # e.g. my_docs/foo/bar.pdf -> output_dir/foo/bar.txt
            rel_path = os.path.relpath(full_path, input_dir)
            base_name = os.path.splitext(rel_path)[0]  # no extension
            txt_rel_path = f"{base_name}.txt"
            out_path = os.path.join(output_dir, txt_rel_path)

            os.makedirs(os.path.dirname(out_path), exist_ok=True)

            raw_text = extract_text_from_file(full_path)
            sanitized_text = sanitize_text(raw_text)
            if not sanitized_text:
                logger.debug("No text extracted (or only control chars) from %s", full_path)
                continue

            # Write to .txt in UTF-8
            try:
                with open(out_path, "w", encoding="utf-8") as out_f:
                    out_f.write(sanitized_text)
                logger.info("Wrote text to %s", out_path)
            except Exception as e:
                logger.error("Error writing to %s: %s", out_path, e)

    logger.info("Conversion complete.")


def zip_output_dir(output_dir, zip_name="converted_files.zip"):
    """
    Zip the entire output directory of .txt files into a single archive.
    """
    zip_path = os.path.join(os.path.dirname(output_dir), zip_name)
    logger.info("Zipping all text files in %s -> %s", output_dir, zip_path)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(output_dir):
            for file_name in files:
                if file_name.endswith(".txt"):
                    full_path = os.path.join(root, file_name)
                    rel_path = os.path.relpath(full_path, output_dir)
                    zf.write(full_path, arcname=rel_path)

    logger.info("Created zip archive at %s", zip_path)


def parse_args():
    parser = argparse.ArgumentParser(description="Convert various filetypes to .txt")
    parser.add_argument("input_dir", help="Directory with original documents")
    parser.add_argument("output_dir", help="Directory to store .txt files")
    parser.add_argument("--zip", nargs="?", const="converted_files.zip", default=None,
                        help="If set, zip the output directory after conversion. Optionally specify a ZIP filename.")
    parser.add_argument("--max-file-size", type=int, default=5 * 1024 * 1024,
                        help="Skip files larger than this (in bytes). Default is 5MB.")
    parser.add_argument("--include-hidden", action="store_true",
                        help="If set, process hidden files & directories.")
    return parser.parse_args()


def main():
    args = parse_args()

    convert_files_to_txt(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        max_file_size=args.max_file_size,
        skip_hidden=not args.include_hidden
    )

    # If requested, zip the resulting .txt files
    if args.zip:
        zip_output_dir(args.output_dir, args.zip)


if __name__ == "__main__":
    main()

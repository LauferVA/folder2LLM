#!/usr/bin/env python3

"""
convert_files_to_txt.py

A standalone script that converts various file formats
(.docx, .pdf, .xlsx, .odt, .ipynb, etc.) to plain-text .txt files.

Usage:
    python convert_files_to_txt.py /path/to/source /path/to/destination

Dependencies:
    - python-docx (for .docx)
    - PyMuPDF (fitz) or PyPDF2 (for .pdf)
    - openpyxl (for .xlsx)
    - odfpy (for .odt)
    - etc.

This script will:
    1) Walk the specified input directory recursively,
    2) Attempt to extract text from recognized file types,
    3) Write that text to a .txt file in the output directory with the same
       relative path and base filename (plus a .txt extension).
    
    e.g., mydoc.pdf -> mydoc.txt
"""

import os
import sys
import logging
import json

# ----------------------------------------------------------------------
# Optional imports for specialized parsing:
# ----------------------------------------------------------------------
try:
    import docx  # For .docx
except ImportError:
    docx = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import openpyxl  # For .xlsx
except ImportError:
    openpyxl = None

try:
    import odf.text, odf.teletype, odf.opendocument  # For .odt
except ImportError:
    odf = None

# ----------------------------------------------------------------------
# Basic logging setup
# ----------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------
# File-parsing routines
# ----------------------------------------------------------------------
def extract_text_from_file(filepath):
    """
    Attempt to extract text from the given filepath, based on extension.
    Return an empty string if no parser is available or if an error occurs.
    """
    ext = os.path.splitext(filepath)[1].lower()

    # Common plaintext or code-like
    if ext in (
        ".txt", ".md", ".py", ".json", ".csv", ".tsv", ".log", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart",
        ".php", ".rb", ".sql"
    ):
        return _read_plaintext(filepath)

    # DOCX
    if ext == ".docx":
        return _read_docx(filepath) if docx else ""

    # .doc or .rtf not fully supported natively
    if ext in (".doc", ".rtf"):
        logger.warning("Native reading not implemented for %s; consider external tools.", ext)
        return ""

    # PDF
    if ext == ".pdf":
        if fitz:
            return _read_pdf_pymupdf(filepath)
        elif PyPDF2:
            return _read_pdf_pypdf2(filepath)
        else:
            logger.warning("No PDF library installed; cannot parse PDFs.")
            return ""

    # ODT
    if ext == ".odt":
        return _read_odt(filepath) if odf else ""

    # Excel
    if ext in (".xlsx", ".xls", ".xlsm", ".ods"):
        return _read_excel(filepath) if openpyxl else ""

    # PowerPoint-like
    if ext in (".ppt", ".pptx", ".odp"):
        logger.warning("Parsing for PPT/ODP not implemented in this script.")
        return ""

    # Jupyter notebook
    if ext == ".ipynb":
        return _read_ipynb(filepath)

    # If no rule matched
    return ""


def _read_plaintext(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        logger.error("Error reading plaintext file %s: %s", filepath, e)
        return ""


def _read_docx(filepath):
    try:
        doc = docx.Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading DOCX %s: %s", filepath, e)
        return ""


def _read_pdf_pymupdf(filepath):
    text_pages = []
    try:
        with fitz.open(filepath) as pdf_doc:
            for page in pdf_doc:
                text_pages.append(page.get_text())
    except Exception as e:
        logger.error("Error reading PDF (PyMuPDF) %s: %s", filepath, e)
    return "\n".join(text_pages)


def _read_pdf_pypdf2(filepath):
    text_pages = []
    try:
        with open(filepath, "rb") as f:
            pdf = PyPDF2.PdfReader(f)
            for page_num in range(len(pdf.pages)):
                text_pages.append(pdf.pages[page_num].extract_text() or "")
    except Exception as e:
        logger.error("Error reading PDF (PyPDF2) %s: %s", filepath, e)
    return "\n".join(text_pages)


def _read_odt(filepath):
    try:
        doc = odf.opendocument.load(filepath)
        text_elements = doc.getElementsByType(odf.text.P)
        paragraphs = [odf.teletype.extractText(elem) for elem in text_elements]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading ODT %s: %s", filepath, e)
        return ""


def _read_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text_chunks = []
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                row_str = "\t".join(str(x) if x is not None else "" for x in row)
                text_chunks.append(row_str)
        return "\n".join(text_chunks)
    except Exception as e:
        logger.error("Error reading Excel file %s: %s", filepath, e)
        return ""


def _read_ipynb(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        text_cells = []
        for cell in data.get("cells", []):
            src = cell.get("source", [])
            text_cells.append("".join(src))
        return "\n".join(text_cells)
    except Exception as e:
        logger.error("Error reading IPYNB %s: %s", filepath, e)
        return ""


# ----------------------------------------------------------------------
# Main conversion logic
# ----------------------------------------------------------------------
def convert_files_to_txt(
    input_dir,
    output_dir,
    max_file_size=5 * 1024 * 1024,
    skip_hidden=True
):
    """
    Recursively scan input_dir for files, extract text, and write to .txt in output_dir.

    :param input_dir: Path to folder containing the original files.
    :param output_dir: Path where converted .txt files will be saved.
    :param max_file_size: Skip files larger than this (in bytes).
    :param skip_hidden: If True, skip hidden files/folders (start with '.').
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    logger.info("Starting conversion from %s to %s", input_dir, output_dir)

    # Example list of recognized extensions
    allowed_extensions = (
        ".txt", ".md", ".py", ".json", ".csv", ".tsv", ".log", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart",
        ".php", ".rb", ".sql", ".doc", ".docx", ".rtf", ".pdf",
        ".odt", ".xls", ".xlsx", ".xlsm", ".ods", ".ppt", ".pptx",
        ".odp", ".ipynb"
    )

    for root, dirs, files in os.walk(input_dir):
        # Optionally skip hidden directories
        if skip_hidden:
            dirs[:] = [d for d in dirs if not d.startswith(".")]

        for file_name in files:
            # Optionally skip hidden files
            if skip_hidden and file_name.startswith("."):
                continue

            ext = os.path.splitext(file_name)[1].lower()
            if ext not in allowed_extensions:
                continue

            full_path = os.path.join(root, file_name)

            # Check file size
            try:
                size = os.path.getsize(full_path)
                if size > max_file_size:
                    logger.debug("Skipping large file %s (size %d bytes)", full_path, size)
                    continue
            except Exception as e:
                logger.error("Could not check file size of %s: %s", full_path, e)
                continue

            # Build the output path
            # e.g., input_dir/foo/bar.pdf -> output_dir/foo/bar.txt
            rel_path = os.path.relpath(full_path, input_dir)
            base_name = os.path.splitext(rel_path)[0]  # drop extension
            txt_rel_path = f"{base_name}.txt"
            out_path = os.path.join(output_dir, txt_rel_path)

            # Ensure directory structure
            os.makedirs(os.path.dirname(out_path), exist_ok=True)

            # Extract text
            text = extract_text_from_file(full_path)
            if not text.strip():
                logger.debug("No text extracted from %s", full_path)
                continue

            # Write to .txt
            try:
                with open(out_path, "w", encoding="utf-8") as out_f:
                    out_f.write(text)
                logger.info("Wrote text to %s", out_path)
            except Exception as e:
                logger.error("Error writing to %s: %s", out_path, e)

    logger.info("Conversion complete.")


def main():
    if len(sys.argv) < 3:
        print("Usage: python convert_files_to_txt.py <input_dir> <output_dir>")
        sys.exit(1)

    input_dir = sys.argv[1]
    output_dir = sys.argv[2]
    convert_files_to_txt(input_dir, output_dir)


if __name__ == "__main__":
    main()

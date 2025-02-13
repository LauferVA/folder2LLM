#!/usr/bin/env python3

"""
convert_files_to_txt.py

Converts various file formats from main script
(.docx, .pdf, .xlsx, .odt, .ipynb, etc.) to cleaned .txt files.

Now enhanced to:
  1) Sanitize and normalize the extracted text (removing odd control chars).
  2) Output everything in UTF-8 encoding.

Usage:
    python convert_files_to_txt.py /path/to/source /path/to/destination
"""

import os
import sys
import logging
import json
import unicodedata

# ----------------------------------------------------------------------
# Optional imports for specialized parsing:
# ----------------------------------------------------------------------
try:
    import docx  # For .docx
except ImportError:
    docx = None

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import openpyxl  # For .xlsx
except ImportError:
    openpyxl = None

try:
    import odf.text, odf.teletype, odf.opendocument  # For .odt
except ImportError:
    odf = None

# ----------------------------------------------------------------------
# Basic logging setup
# ----------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def sanitize_text(text):
    """
    Remove or replace odd/control characters and ensure UTF-8 friendly content.
    - Normalize unicode to NFC form.
    - Optionally remove control characters except for newlines, tabs, etc.

    Customize this function further if you wish to strip all non-ASCII or handle other edge cases.
    """
    # 1) Normalize Unicode to a canonical form (NFC).
    text = unicodedata.normalize("NFC", text)

    # 2) Remove non-printable control characters, but keep newlines/tabs.
    #    'isprintable()' returns False for a number of unicode chars, including some emojis.
    #    If you want to keep extended Unicode (like emojis), you might skip this approach
    #    or do something more targeted. For demonstration, let's allow standard newlines
    #    and tabs, but remove everything else that is non-printable.
    cleaned_chars = []
    for ch in text:
        # If it's a standard printable char or a newline/tab, keep it
        if ch.isprintable() or ch in ['\n', '\r', '\t']:
            cleaned_chars.append(ch)
        else:
            # Replace unprintable chars with a space (or remove them entirely)
            cleaned_chars.append(" ")

    sanitized = "".join(cleaned_chars)

    # 3) Trim leading/trailing whitespace
    sanitized = sanitized.strip()

    return sanitized


# ----------------------------------------------------------------------
# File-parsing routines
# ----------------------------------------------------------------------
def extract_text_from_file(filepath):
    """
    Attempt to extract text from the given filepath, based on extension.
    Return an empty string if no parser is available or if an error occurs.
    """
    ext = os.path.splitext(filepath)[1].lower()

    # Common plaintext or code-like
    if ext in (
        ".txt", ".md", ".py", ".json", ".csv", ".tsv", ".log", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart",
        ".php", ".rb", ".sql"
    ):
        return _read_plaintext(filepath)

    # DOCX
    if ext == ".docx":
        return _read_docx(filepath) if docx else ""

    # .doc or .rtf not fully supported natively
    if ext in (".doc", ".rtf"):
        logger.warning("Native reading not implemented for %s; consider external tools.", ext)
        return ""

    # PDF
    if ext == ".pdf":
        if fitz:
            return _read_pdf_pymupdf(filepath)
        elif PyPDF2:
            return _read_pdf_pypdf2(filepath)
        else:
            logger.warning("No PDF library installed; cannot parse PDFs.")
            return ""

    # ODT
    if ext == ".odt":
        return _read_odt(filepath) if odf else ""

    # Excel
    if ext in (".xlsx", ".xls", ".xlsm", ".ods"):
        return _read_excel(filepath) if openpyxl else ""

    # PowerPoint-like
    if ext in (".ppt", ".pptx", ".odp"):
        logger.warning("Parsing for PPT/ODP not implemented in this script.")
        return ""

    # Jupyter notebook
    if ext == ".ipynb":
        return _read_ipynb(filepath)

    # If no rule matched
    return ""


def _read_plaintext(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        logger.error("Error reading plaintext file %s: %s", filepath, e)
        return ""


def _read_docx(filepath):
    try:
        doc = docx.Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading DOCX %s: %s", filepath, e)
        return ""


def _read_pdf_pymupdf(filepath):
    text_pages = []
    try:
        with fitz.open(filepath) as pdf_doc:
            for page in pdf_doc:
                text_pages.append(page.get_text())
    except Exception as e:
        logger.error("Error reading PDF (PyMuPDF) %s: %s", filepath, e)
    return "\n".join(text_pages)


def _read_pdf_pypdf2(filepath):
    text_pages = []
    try:
        with open(filepath, "rb") as f:
            pdf = PyPDF2.PdfReader(f)
            for page_num in range(len(pdf.pages)):
                text_pages.append(pdf.pages[page_num].extract_text() or "")
    except Exception as e:
        logger.error("Error reading PDF (PyPDF2) %s: %s", filepath, e)
    return "\n".join(text_pages)


def _read_odt(filepath):
    try:
        doc = odf.opendocument.load(filepath)
        text_elements = doc.getElementsByType(odf.text.P)
        paragraphs = [odf.teletype.extractText(elem) for elem in text_elements]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading ODT %s: %s", filepath, e)
        return ""


def _read_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text_chunks = []
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                row_str = "\t".join(str(x) if x is not None else "" for x in row)
                text_chunks.append(row_str)
        return "\n".join(text_chunks)
    except Exception as e:
        logger.error("Error reading Excel file %s: %s", filepath, e)
        return ""


def _read_ipynb(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        text_cells = []
        for cell in data.get("cells", []):
            src = cell.get("source", [])
            text_cells.append("".join(src))
        return "\n".join(text_cells)
    except Exception as e:
        logger.error("Error reading IPYNB %s: %s", filepath, e)
        return ""


# ----------------------------------------------------------------------
# Main conversion logic
# ----------------------------------------------------------------------
def convert_files_to_txt(
    input_dir,
    output_dir,
    max_file_size=5 * 1024 * 1024,
    skip_hidden=True
):
    """
    Recursively scan input_dir for files, extract text, sanitize it,
    then write to .txt in output_dir.

    :param input_dir: Path to folder containing the original files.
    :param output_dir: Path where converted .txt files will be saved.
    :param max_file_size: Skip files larger than this (in bytes).
    :param skip_hidden: If True, skip hidden files/folders (start with '.').
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    logger.info("Starting conversion from %s to %s", input_dir, output_dir)

    # Example list of recognized extensions
    allowed_extensions = (
        ".txt", ".md", ".py", ".json", ".csv", ".tsv", ".log", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart",
        ".php", ".rb", ".sql", ".doc", ".docx", ".rtf", ".pdf",
        ".odt", ".xls", ".xlsx", ".xlsm", ".ods", ".ppt", ".pptx",
        ".odp", ".ipynb"
    )

    for root, dirs, files in os.walk(input_dir):
        # Optionally skip hidden directories
        if skip_hidden:
            dirs[:] = [d for d in dirs if not d.startswith(".")]

        for file_name in files:
            # Optionally skip hidden files
            if skip_hidden and file_name.startswith("."):
                continue

            ext = os.path.splitext(file_name)[1].lower()
            if ext not in allowed_extensions:
                continue

            full_path = os.path.join(root, file_name)

            # Check file size
            try:
                size = os.path.getsize(full_path)
                if size > max_file_size:
                    logger.debug("Skipping large file %s (size %d bytes)", full_path, size)
                    continue
            except Exception as e:
                logger.error("Could not check file size of %s: %s", full_path, e)
                continue

            # Build the output path
            # e.g., input_dir/foo/bar.pdf -> output_dir/foo/bar.txt
            rel_path = os.path.relpath(full_path, input_dir)
            base_name = os.path.splitext(rel_path)[0]  # drop extension
            txt_rel_path = f"{base_name}.txt"
            out_path = os.path.join(output_dir, txt_rel_path)

            # Ensure directory structure
            os.makedirs(os.path.dirname(out_path), exist_ok=True)

            # Extract text
            raw_text = extract_text_from_file(full_path)
            sanitized_text = sanitize_text(raw_text)
            if not sanitized_text:
                logger.debug("No text extracted (or only control chars) from %s", full_path)
                continue

            # Write to .txt in UTF-8
            try:
                with open(out_path, "w", encoding="utf-8") as out_f:
                    out_f.write(sanitized_text)
                logger.info("Wrote text to %s", out_path)
            except Exception as e:
                logger.error("Error writing to %s: %s", out_path, e)

    logger.info("Conversion complete.")


def main():
    if len(sys.argv) < 3:
        print("Usage: python convert_files_to_txt.py <input_dir> <output_dir>")
        sys.exit(1)

    input_dir = sys.argv[1]
    output_dir = sys.argv[2]
    convert_files_to_txt(input_dir, output_dir)


if __name__ == "__main__":
    main()
